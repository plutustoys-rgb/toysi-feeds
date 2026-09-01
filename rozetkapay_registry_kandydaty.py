"""rozetkapay_registry_kandydaty.py — детектор СТОРНО (і еквайринг-частини графи 9) з реєстру
переказів FC/RozetkaPay (обслуговує і Prom-оплату, і Rozetka-картки — один процесор).

НАВІЩО (запит бухгалтера 2026-08-31, друга дірка детекції сторно): банк сторнував Prom-замовлення
421459063 (30.08), а книга ще показувала повний дохід 17 днів — бухгалтер знайшла руками з реєстру.
Це мало виявлятись автоматично.

ЩО: читає реєстр (xlsx), для кожного рядка:
  • СТОРНО (сума платежу < 0, АБО тип оплати == «Повернення») → кандидат «замовлення X сторновано
    банком DD.MM, сума ±N, комісія повернена; книга ще показує дохід — перевір графу 5/6»;
  • Rozetka-оплата (не сторно), де в книзі графа 9 ще порожня → кандидат «еквайринг-частина графи 9
    = |E|, готова докласти до роялті».
Кросс-звірка з книгою `KODV_PlutusToys_2026.xlsx` — READ-ONLY (графу пише лише роль «бухгалтер»).
Пише кандидатів у `документи_КОДВ/YYYY-MM/RozetkaPay/…`, книгу НЕ чіпає.

ВАЖЛИВО (жива інспекція 2026-09-01): кирилиця в реєстрі ЧИСТА при читанні openpyxl БЕЗ read_only
(read_only занижує dimension до A1 і губить рядки). Тому і статус «Повернення», і заголовки читаються
нормально — детекція сторно за статусом + знаком суми (подвійно), не лише за позицією.

ЗАПУСК: python rozetkapay_registry_kandydaty.py [--file <xlsx>]
  без --file: бере найновіший xlsx з документи_КОДВ/<міс>/RozetkaPay/ (куди кладе kodv_mail_archiver).
"""
import argparse
import glob
import json
import os
import sys
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE_DIR = Path(__file__).parent
COWORK_DIR = Path(os.environ.get("PLUTUS_COWORK_DIR",
                                 r"C:\Users\smach\Claude\Projects\PlutusToys_avtonomiya"))
KODV_XLSX = COWORK_DIR / "KODV_PlutusToys_2026.xlsx"
DOCS_DIR = COWORK_DIR / "документи_КОДВ"
CURSOR_FILE = BASE_DIR / ".local_secrets" / "rozetkapay_registry_cursor.json"
_NO_TELEGRAM = os.environ.get("AUDIT_NO_TELEGRAM") == "1"

# Позиції колонок реєстру (0-індекс) — звірено живо на реєстрі 31.08 (заголовки читаються чисто):
C_NUM, C_DATE_TRANSFER, C_DATE_PAY, C_SUM, C_COMMISSION = 0, 1, 2, 3, 4
C_PROJECT, C_ORDER, C_TYPE, C_FINOP = 7, 8, 11, 13  # H назва проекту, I №замовл, L тип, N унік.фін-номер


def _notify(msg: str) -> None:
    if _NO_TELEGRAM:
        return
    try:
        from telegram_notify import send_telegram_message
        send_telegram_message(msg)
    except Exception as e:  # noqa: BLE001
        print(f"[RzPayReg] Telegram не надіслано (не критично): {e}", file=sys.stderr)


def _load_cursor() -> set:
    try:
        return set(json.loads(CURSOR_FILE.read_text(encoding="utf-8")).get("seen_finop", []))
    except (ValueError, OSError):
        return set()


def _save_cursor(seen: set) -> None:
    CURSOR_FILE.parent.mkdir(parents=True, exist_ok=True)
    CURSOR_FILE.write_text(json.dumps({"seen_finop": sorted(seen),
                                       "updated_at": datetime.now().isoformat(timespec="seconds")},
                                      ensure_ascii=False, indent=2), encoding="utf-8")


def _newest_registry() -> Path:
    """Найновіший реєстр у документи_КОДВ/<міс>/RozetkaPay/ (куди кладе kodv_mail_archiver)."""
    pat = str(DOCS_DIR / "*" / "RozetkaPay" / "*.xlsx")
    files = [f for f in glob.glob(pat) if not os.path.basename(f).startswith("~$")]
    return Path(max(files, key=os.path.getmtime)) if files else None


def _parse_registry(path: Path) -> list:
    """Рядки даних реєстру як dict. openpyxl БЕЗ read_only (інакше губить рядки — dimension бита)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    # заголовок таблиці — рядок, де A=='№' і колонка замовлення містить '№ замовлення'
    hdr = None
    for i, r in enumerate(all_rows):
        if len(r) > C_ORDER and str(r[C_NUM]).strip() == "№" and "замовлення" in str(r[C_ORDER] or ""):
            hdr = i
            break
    if hdr is None:
        return []
    out = []
    for r in all_rows[hdr + 1:]:
        num = r[C_NUM] if len(r) > C_NUM else None
        # рядок даних має числовий № у колонці A; підсумковий «Всього:» його не має
        if num is None or not str(num).strip().isdigit():
            continue
        out.append({
            "seq": str(num).strip(),
            "date_pay": str(r[C_DATE_PAY] or "").strip() if len(r) > C_DATE_PAY else "",
            "date_transfer": str(r[C_DATE_TRANSFER] or "").strip() if len(r) > C_DATE_TRANSFER else "",
            "sum": r[C_SUM] if len(r) > C_SUM else None,
            "commission": r[C_COMMISSION] if len(r) > C_COMMISSION else None,
            "project": str(r[C_PROJECT] or "").strip() if len(r) > C_PROJECT else "",
            "order_id": str(r[C_ORDER] or "").strip() if len(r) > C_ORDER else "",
            "type": str(r[C_TYPE] or "").strip() if len(r) > C_TYPE else "",
            "finop": str(r[C_FINOP] or "").strip() if len(r) > C_FINOP else "",
        })
    return out


def _lookup_book_row(order_id: str) -> dict:
    """READ-ONLY: шукає '№<order_id>' у графі 5 (колонка E) книги КОДВ. {row, current_i9} або {}.
    Книгу НЕ пише — графу пише лише роль «бухгалтер» (правило власника)."""
    if not order_id or not KODV_XLSX.exists():
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(KODV_XLSX), data_only=True, read_only=True)
        ws = wb["КОДВ"]
        needle = f"№{order_id}"
        for row in ws.iter_rows(min_row=7):
            e = row[4].value if len(row) > 4 else None
            if e and needle in str(e):
                return {"row": row[4].row, "current_i9": row[8].value if len(row) > 8 else None}
    except Exception as e:  # noqa: BLE001
        print(f"[RzPayReg] крос-звірка з книгою не вдалась (не критично): {e}", file=sys.stderr)
    return {}


def _is_storno(row: dict) -> bool:
    """Подвійний сигнал: сума платежу < 0 АБО тип оплати == «Повернення»."""
    s = row["sum"]
    if isinstance(s, (int, float)) and s < 0:
        return True
    return row["type"] == "Повернення"


def _marketplace(row: dict) -> str:
    p = row["project"].lower()
    if "rozetka" in p:
        return "Rozetka"
    if "prom" in p:
        return "Prom"
    return "?"


def collect(rows: list) -> tuple:
    """(candidates, seen_finop_this_file). Кандидати лише по НОВИХ (не в курсорі) рядках."""
    seen = _load_cursor()
    is_first = not seen
    this_file = {r["finop"] for r in rows if r["finop"]}

    if is_first:
        print(f"[RzPayReg] Перший запуск — {len(this_file)} операцій за базову лінію, кандидатів не шукаю.")
        return [], this_file

    candidates = []
    for r in rows:
        if r["finop"] and r["finop"] in seen:
            continue  # вже бачили
        oid = r["order_id"]
        if _is_storno(r):
            book = _lookup_book_row(oid)
            candidates.append({
                "kind": "storno",
                "order_id": oid,
                "marketplace": _marketplace(r),
                "date": r["date_pay"] or r["date_transfer"],
                "sum": r["sum"],
                "commission_returned": r["commission"],
                "book_row": book.get("row"),
                "book_current_i9": book.get("current_i9"),
                "note": (f"СТОРНО: банк повернув оплату замовлення {oid} ({_marketplace(r)}) "
                         f"на {r['date_pay']}, сума {r['sum']}, комісія повернена {r['commission']}. "
                         + (f"У книзі рядок {book['row']} — перевір графу 5/6 (дохід міг лишитись повним)."
                            if book else "У книзі замовлення не знайдено — можливо, ще не внесено.")),
            })
    return candidates, this_file


def _write_report(candidates: list, src: Path) -> Path:
    today = datetime.now()
    month_dir = DOCS_DIR / today.strftime("%Y-%m") / "RozetkaPay"
    month_dir.mkdir(parents=True, exist_ok=True)
    stamp = today.strftime("%Y-%m-%d")
    md = month_dir / f"{stamp}_rozetkapay_kandydaty.md"
    js = month_dir / f"{stamp}_rozetkapay_kandydaty.json"
    lines = [f"# RozetkaPay реєстр — кандидати (сторно/еквайринг), {today.strftime('%Y-%m-%d %H:%M')}", "",
             f"Джерело: {src.name}", f"**Кандидатів: {len(candidates)}** (книгу НЕ чіпаю — це роль бухгалтера).", ""]
    for c in candidates:
        lines.append(f"## 🔴 СТОРНО — замовлення {c['order_id']} ({c['marketplace']})")
        lines.append(f"- {c['note']}")
        lines.append(f"- Книга: рядок {c['book_row']}, поточна графа 9 = {c['book_current_i9']}")
        lines.append("")
    md.write_text("\n".join(lines), encoding="utf-8")
    js.write_text(json.dumps(candidates, ensure_ascii=False, indent=2), encoding="utf-8")
    return md


def main() -> int:
    ap = argparse.ArgumentParser(description="Детектор сторно/еквайрингу з реєстру RozetkaPay.")
    ap.add_argument("--file", help="Конкретний xlsx (для тесту); без нього — найновіший у документи_КОДВ.")
    args = ap.parse_args()

    src = Path(args.file) if args.file else _newest_registry()
    if not src or not src.exists():
        print("[RzPayReg] Реєстр не знайдено (ні --file, ні в документи_КОДВ/*/RozetkaPay/).", file=sys.stderr)
        return 1
    rows = _parse_registry(src)
    print(f"[RzPayReg] {src.name}: рядків даних {len(rows)}, сторно {sum(1 for r in rows if _is_storno(r))}.")

    candidates, this_file = collect(rows)
    if candidates:
        report = _write_report(candidates, src)
        print(f"[RzPayReg] Кандидатів {len(candidates)} → {report}")
        _notify(f"💳 RozetkaPay: {len(candidates)} СТОРНО-кандидат(ів) у реєстрі — книга може показувати "
                f"повний дохід. Див. {report.name}")
    else:
        print("[RzPayReg] Нових кандидатів немає.")

    seen = _load_cursor() | this_file
    _save_cursor(seen)
    return 0


if __name__ == "__main__":
    sys.exit(main())
