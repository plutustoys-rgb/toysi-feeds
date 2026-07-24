"""
novapay_statement.py — автоматична звірка надходжень NovaPay проти orders_db.

АРХІТЕКТУРА (2026-07-24, пряме прохання власниці, після живого дослідження)
Прямий SOAP API (novapay_client.py, UserAuthenticationJWT) виявився заблокованим
NullReferenceException на боці NovaPay (питання відправлено в підтримку,
відповіді ще немає) — і навіть якби працював, GetPaymentsList дає лише
АГРЕГОВАНІ суми по реєстру, ніколи по кожному замовленню.

Реальний робочий шлях, знайдений через живий приклад файлу від власниці:
NovaPay САМА надсилає "Реєстр переказів №XXXXX" (.XLSX) на email власниці
автоматично (не ручний експорт) — і, що критично, колонка "Номер замовлення"
в цьому реєстрі вже містить НАШ internal_order_id напряму
("prom_416236076 / 10044575" — наш internal_order_id / toysi_order_id), тож
зіставлення з orders_db точне, без вгадування. Це покриває САМЕ COD
(готівкове зарахування Нової Пошти) — колонка "№ ЕН НП" (номер ТТН)
підтверджує це для кожного рядка.

Тому: IMAP-опитування поштової скриньки (NOVAPAY_IMAP_EMAIL/
NOVAPAY_IMAP_APP_PASSWORD — вже налаштовані й живо перевірені на VPS з
16-17.07), не SOAP API. novapay_client.py лишається в репозиторії
(може стати в пригоді, якщо колись знадобиться щось за межами вже
надісланих реєстрів), але НЕ використовується цим скриптом.

Механізм:
1. IMAP UNSEEN-повідомлення зі вкладенням .xlsx, ім'я якого містить "Реєстр"
   (не залежимо від точної адреси відправника/теми — стійкіше до змін).
2. Парсимо кожен реєстр (openpyxl) — з обхідним рішенням для реального бага
   експорту NovaPay: вкладення має xl/SharedStrings.xml (велика "S"), а не
   стандартне xl/sharedStrings.xml, тому звичайний openpyxl.load_workbook()
   падає з KeyError. Пересобираємо zip з виправленим ім'ям файлу в пам'яті.
3. Кожен рядок реєстру -> internal_order_id (частина "Номер замовлення" до
   " / ") -> шукаємо в orders_db. Записуємо результат (знайдено чи ні) —
   НЕ вигадуємо відповідність, якщо order_id не знайдено, лишаємо позначку
   "не знайдено" в журналі, а не тихо пропускаємо.
4. Дописуємо в kodv_ledger.jsonl (той самий файл, що й daily_report.py,
   новий тип запису "novapay_reconciliation" — не конфліктує з існуючими
   per-order записами daily_report.py, лінійний журнал, легко відрізнити
   за полем "type").
5. Позначаємо email прочитаним (\\Seen) ЛИШЕ після успішного запису в
   журнал — якщо скрипт впаде посередині, лист лишиться непрочитаним і
   безпечно обробиться повторно наступного запуску (той самий принцип, що
   й escalate() у prom_chat_bot.py).
6. Додатковий, незалежний захист від дублів — processed-registries.json
   зі списком уже оброблених номерів реєстру (на випадок повторної
   відправки того самого листа чи ручного повторного запуску).

Запуск:
    python novapay_statement.py
"""
import imaplib
import email
import io
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timezone
from email.header import decode_header
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from daily_report import KODV_LEDGER_FILE
from orders_db import get_connection, get_order
from telegram_notify import send_telegram_message

load_dotenv()

NOVAPAY_IMAP_EMAIL = os.environ.get("NOVAPAY_IMAP_EMAIL", "")
NOVAPAY_IMAP_APP_PASSWORD = os.environ.get("NOVAPAY_IMAP_APP_PASSWORD", "")
# Gmail — найпоширеніший провайдер з "App Password" (термін самої NovaPay-
# інструкції з pt20 збігається з Gmail-термінологією); перевизначувано
# через .env, якщо скринька насправді на іншому провайдері.
NOVAPAY_IMAP_HOST = os.environ.get("NOVAPAY_IMAP_HOST", "imap.gmail.com")

BASE_DIR = Path(__file__).parent
PROCESSED_REGISTRIES_FILE = BASE_DIR / "novapay_processed_registries.json"

# Ім'я вкладення реєстру завжди містить це слово (перевірено на реальному
# зразку: "Реєстр платежів контрагента ... .XLSX") — не залежимо від точної
# адреси відправника чи теми листа, стійкіше до майбутніх змін формулювання
# NovaPay.
_ATTACHMENT_NAME_MARKER = "реєстр"


class NovaPayStatementError(Exception):
    pass


def _load_processed_registries() -> set:
    if not PROCESSED_REGISTRIES_FILE.exists():
        return set()
    try:
        return set(json.loads(PROCESSED_REGISTRIES_FILE.read_text(encoding="utf-8")))
    except (json.JSONDecodeError, ValueError):
        return set()


def _save_processed_registries(registries: set) -> None:
    PROCESSED_REGISTRIES_FILE.write_text(
        json.dumps(sorted(registries), ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _decode_mime_words(raw: str) -> str:
    decoded_parts = decode_header(raw)
    return "".join(
        part.decode(encoding or "utf-8", errors="replace") if isinstance(part, bytes) else part
        for part, encoding in decoded_parts
    )


def _fetch_unseen_registry_attachments(imap_conn: imaplib.IMAP4_SSL) -> list:
    """Повертає список (uid, filename, bytes) для непрочитаних листів із
    вкладенням .xlsx, ім'я якого містить _ATTACHMENT_NAME_MARKER. Листи БЕЗ
    такого вкладення НЕ чіпаємо (не позначаємо прочитаними) — вони можуть
    бути про щось інше, не наша справа їх чіпати."""
    imap_conn.select("INBOX")
    status, data = imap_conn.search(None, "UNSEEN")
    if status != "OK":
        raise NovaPayStatementError(f"IMAP SEARCH не вдався: {status}")

    uids = data[0].split()
    results = []
    for uid in uids:
        status, msg_data = imap_conn.fetch(uid, "(RFC822)")
        if status != "OK" or not msg_data or not msg_data[0]:
            continue
        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)

        for part in msg.walk():
            filename = part.get_filename()
            if not filename:
                continue
            filename = _decode_mime_words(filename)
            if not filename.lower().endswith(".xlsx"):
                continue
            if _ATTACHMENT_NAME_MARKER not in filename.lower():
                continue
            payload = part.get_payload(decode=True)
            if payload:
                results.append((uid, filename, payload))
            break  # одне вкладення-реєстр на лист, немає сенсу шукати далі
    return results


def _mark_seen(imap_conn: imaplib.IMAP4_SSL, uid: bytes) -> None:
    imap_conn.store(uid, "+FLAGS", "\\Seen")


def _fix_shared_strings_casing(xlsx_bytes: bytes) -> bytes:
    """Реальний, підтверджений баг експорту NovaPay: вкладення має
    xl/SharedStrings.xml (велика 'S'), а не стандартне xl/sharedStrings.xml
    з специфікації OOXML — openpyxl шукає рівно другий варіант і падає з
    KeyError без цього обхідного рішення. Пересобираємо zip в пам'яті лише
    з виправленим ім'ям одного файлу, решта байтів незмінні."""
    buf_in = io.BytesIO(xlsx_bytes)
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_in) as zin, zipfile.ZipFile(buf_out, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            name = "xl/sharedStrings.xml" if item.filename == "xl/SharedStrings.xml" else item.filename
            zout.writestr(name, data)
    return buf_out.getvalue()


_REGISTRY_NUMBER_RE = re.compile(r"РЕЄСТР ПЕРЕКАЗІВ\s*№\s*(\d+)")

# Не жорстко прив'язуємось до індексу колонки (порядок може змінитись у
# майбутньому експорті NovaPay) — шукаємо за підрядком заголовка, стійкіше.
_COLUMN_MARKERS = {
    "date": "дата перерахунку",
    "amount_received": "сума принятих",
    "commission": "сума утриманої",
    "amount_net": "сума перерахованих",
    "buyer_name": "піб покупця",
    "ttn": "ен нп",
    "order_ref": "номер замовлення",
}


def parse_registry_xlsx(xlsx_bytes: bytes) -> dict:
    """Повертає {"registry_number": str, "rows": [ {date, amount_received,
    commission, amount_net, buyer_name, ttn, internal_order_id,
    toysi_order_id, raw_order_ref}, ... ]}. Рядки без розпізнаваного
    internal_order_id (formатy не "prom_.../rozetka_..." тощо) все одно
    повертаються — з internal_order_id=None, щоб їх було видно в журналі
    як "не зіставлено", а не тихо загублено."""
    fixed_bytes = _fix_shared_strings_casing(xlsx_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(fixed_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    registry_number = None
    header_row_idx = None
    column_index = {}

    all_rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(all_rows):
        for cell in row:
            if isinstance(cell, str):
                match = _REGISTRY_NUMBER_RE.search(cell)
                if match:
                    registry_number = match.group(1)

        if header_row_idx is None:
            normalized_cells = [(cell or "").strip().lower() if isinstance(cell, str) else "" for cell in row]
            found = {}
            for key, marker in _COLUMN_MARKERS.items():
                for col_idx, cell_text in enumerate(normalized_cells):
                    if marker in cell_text:
                        found[key] = col_idx
                        break
            if len(found) == len(_COLUMN_MARKERS):
                header_row_idx = i
                column_index = found

    if registry_number is None:
        raise NovaPayStatementError("Не знайдено номер реєстру (РЕЄСТР ПЕРЕКАЗІВ №...) у файлі")
    if header_row_idx is None:
        raise NovaPayStatementError("Не знайдено рядок заголовків очікуваної структури у файлі")

    rows = []
    for row in all_rows[header_row_idx + 1:]:
        order_ref_cell = row[column_index["order_ref"]] if column_index["order_ref"] < len(row) else None
        if not isinstance(order_ref_cell, str) or not order_ref_cell.strip():
            continue  # порожній/підсумковий рядок ("Всього:") — не запис про переказ

        raw_order_ref = order_ref_cell.strip()
        internal_order_id, toysi_order_id = None, None
        if "/" in raw_order_ref:
            left, _, right = raw_order_ref.partition("/")
            internal_order_id = left.strip()
            toysi_order_id = right.strip()
        else:
            internal_order_id = raw_order_ref

        def _cell(key):
            idx = column_index[key]
            return row[idx] if idx < len(row) else None

        rows.append({
            "date": _cell("date"),
            "amount_received": _cell("amount_received"),
            "commission": _cell("commission"),
            "amount_net": _cell("amount_net"),
            "buyer_name": (_cell("buyer_name") or "").strip() if isinstance(_cell("buyer_name"), str) else _cell("buyer_name"),
            "ttn": _cell("ttn"),
            "raw_order_ref": raw_order_ref,
            "internal_order_id": internal_order_id,
            "toysi_order_id": toysi_order_id,
        })

    return {"registry_number": registry_number, "rows": rows}


def _match_order(conn, internal_order_id: str | None) -> dict | None:
    if not internal_order_id:
        return None
    return get_order(conn, internal_order_id)


def _append_ledger_entries(registry_number: str, rows: list, matched_flags: list) -> None:
    now_iso = datetime.now(timezone.utc).isoformat()
    lines = []
    for row, matched in zip(rows, matched_flags):
        entry = {
            "type": "novapay_reconciliation",
            "logged_at": now_iso,
            "registry_number": registry_number,
            "date": row["date"],
            "internal_order_id": row["internal_order_id"],
            "toysi_order_id": row["toysi_order_id"],
            "raw_order_ref": row["raw_order_ref"],
            "ttn": row["ttn"],
            "buyer_name": row["buyer_name"],
            "amount_received": row["amount_received"],
            "commission": row["commission"],
            "amount_net": row["amount_net"],
            "matched_in_orders_db": matched,
        }
        lines.append(json.dumps(entry, ensure_ascii=False))
    if lines:
        with KODV_LEDGER_FILE.open("a", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")


def process_registry_email(conn, uid: bytes, filename: str, xlsx_bytes: bytes, processed: set) -> dict:
    parsed = parse_registry_xlsx(xlsx_bytes)
    registry_number = parsed["registry_number"]

    if registry_number in processed:
        print(f"[NovaPay] Реєстр №{registry_number} ({filename}) вже оброблено раніше — пропускаю, лише позначаю прочитаним.")
        return {"registry_number": registry_number, "rows": 0, "matched": 0, "unmatched": 0, "skipped_duplicate": True}

    matched_flags = []
    matched_count = 0
    for row in parsed["rows"]:
        order = _match_order(conn, row["internal_order_id"])
        is_matched = order is not None
        matched_flags.append(is_matched)
        if is_matched:
            matched_count += 1

    _append_ledger_entries(registry_number, parsed["rows"], matched_flags)
    processed.add(registry_number)

    return {
        "registry_number": registry_number,
        "rows": len(parsed["rows"]),
        "matched": matched_count,
        "unmatched": len(parsed["rows"]) - matched_count,
        "skipped_duplicate": False,
    }


def main() -> None:
    if not (NOVAPAY_IMAP_EMAIL and NOVAPAY_IMAP_APP_PASSWORD):
        print("[NovaPay] NOVAPAY_IMAP_EMAIL/NOVAPAY_IMAP_APP_PASSWORD не задані в .env — зупиняюсь.", file=sys.stderr)
        sys.exit(1)

    processed = _load_processed_registries()

    try:
        imap_conn = imaplib.IMAP4_SSL(NOVAPAY_IMAP_HOST)
        imap_conn.login(NOVAPAY_IMAP_EMAIL, NOVAPAY_IMAP_APP_PASSWORD)
    except (imaplib.IMAP4.error, OSError) as e:
        message = f"🚨 novapay_statement.py: не вдалось підключитись до IMAP ({NOVAPAY_IMAP_HOST}): {e}"
        print(f"[NovaPay] {message}", file=sys.stderr)
        send_telegram_message(message)
        sys.exit(1)

    try:
        attachments = _fetch_unseen_registry_attachments(imap_conn)
        print(f"[NovaPay] Непрочитаних листів із реєстром: {len(attachments)}.")

        if not attachments:
            return

        results = []
        with get_connection() as conn:
            for uid, filename, xlsx_bytes in attachments:
                try:
                    result = process_registry_email(conn, uid, filename, xlsx_bytes, processed)
                    results.append(result)
                    _mark_seen(imap_conn, uid)
                    print(f"[NovaPay] Реєстр №{result['registry_number']}: {result['rows']} рядків, "
                          f"{result['matched']} зіставлено, {result['unmatched']} без відповідника в orders_db.")
                except NovaPayStatementError as e:
                    message = (
                        f"⚠️ novapay_statement.py: не вдалось розпарсити вкладення '{filename}': {e} — "
                        f"лист лишається непрочитаним, повторна спроба наступного запуску."
                    )
                    print(f"[NovaPay] {message}", file=sys.stderr)
                    send_telegram_message(message)
                except Exception as e:
                    # ВИПРАВЛЕНО (аудит PR #156, pt8): будь-яка НЕОЧІКУВАНА помилка тут
                    # раніше пропагувалась вище й аварійно завершувала весь скрипт БЕЗ
                    # жодного сповіщення — для скрипта, чия єдина мета фінансова звірка,
                    # тиха відмова означає пропущену звірку, яку ніхто не помітить. Той
                    # самий принцип, що й у order_status_tracker.py/service_watchdog.py.
                    message = (
                        f"🚨 novapay_statement.py: неочікувана помилка обробки вкладення "
                        f"'{filename}': {e} — лист лишається непрочитаним для повторної спроби."
                    )
                    print(f"[NovaPay] {message}", file=sys.stderr)
                    send_telegram_message(message)

        _save_processed_registries(processed)

        new_results = [r for r in results if not r["skipped_duplicate"]]
        if new_results:
            total_rows = sum(r["rows"] for r in new_results)
            total_matched = sum(r["matched"] for r in new_results)
            total_unmatched = sum(r["unmatched"] for r in new_results)
            summary = (
                f"💳 NovaPay-звірка: {len(new_results)} нових реєстрів, {total_rows} переказів, "
                f"{total_matched} зіставлено з orders_db"
                + (f", {total_unmatched} без відповідника" if total_unmatched else "")
                + "."
            )
            send_telegram_message(summary)
            print(f"[NovaPay] {summary}")
    except Exception as e:
        # Той самий принцип, що й вище — будь-яка помилка ПОЗА циклом обробки
        # окремих вкладень (наприклад, сам IMAP SEARCH чи запис processed-
        # registries на диск) теж повинна дійти до Telegram, не лише traceback
        # у логах VPS, які ніхто не читає без окремого приводу.
        message = f"🚨 novapay_statement.py: неочікувана помилка: {e}"
        print(f"[NovaPay] {message}", file=sys.stderr)
        send_telegram_message(message)
        raise
    finally:
        try:
            imap_conn.logout()
        except Exception:
            pass


if __name__ == "__main__":
    main()
